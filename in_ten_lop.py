from openpyxl import load_workbook, Workbook
workbook = load_workbook(filename="lop.xlsx")
sheet_1 = workbook["Sheet1"]

mr = sheet_1.max_row
mc = sheet_1.max_column

for i in range(2,mr+1):
    if sheet_1.cell(row=i,column=2).value>5:
        print(sheet_1.cell(row=i,column=1).value)
workbook.close()
