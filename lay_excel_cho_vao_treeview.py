from tkinter import *
from tkinter.ttk import Treeview
from openpyxl import *

#vmt
# my_form = Tk()
#
# ds_cot = []
# ds_data = []
# ds_data_chung = []
# wb1 = load_workbook(filename="dulieu.xlsx")
# ws1 = wb1.worksheets[0]
#
# mr = ws1.max_row
# mc = ws1.max_column
# # for i in range(1, mr + 1):
# #     for j in range(1, mc + 1):
# #         # Vị trí ô chứ dữ liệu trong ws1
# #         c = ws1.cell(row=i, column=j)
#
# for i in range(1, mc + 1):
#     cell_cot = ws1.cell(row=1, column=i).value
#     ds_cot.append(cell_cot)
# for j in range(2, mr + 1):
#     if len(ds_data) != 0:
#         ds_data_chung.append(tuple(ds_data))
#     ds_data = []
#     for i in range(1, mc + 1):
#         cell_data = ws1.cell(row=j, column=i).value
#         ds_data.append(cell_data)
#
# # ds_du_lieu = [('data1', 'data2', 'data3'), ('data4', 'data5', 'data6'), ('data7', 'data8', 'data8')]
# # ds_cot = ('cột 1', 'Cột 2', 'Cột 3') #ok
#
# listbox = Treeview(master=my_form, columns=ds_cot, show='headings')
# for cot in ds_cot:
#     listbox.heading(cot, text=cot)
#     listbox.column(cot, width=120)
#
# du_lieu_trong_cot = ds_data_chung
# for cot1, cot2, cot3,cot4 in du_lieu_trong_cot:
#     listbox.insert("", "end", values=(cot1, cot2, cot3,cot4))
#
# listbox.pack()
#
# wb1.close()
# my_form.mainloop()

#thay
def get_data_excel():
    # Mở file excel
    lst_data = []
    workbook = load_workbook(filename="dulieu.xlsx")
    sheet = workbook["Sheet1"]

# Doc du lieu trona file excel d o Al
    for i in range(2, sheet.max.row + 1):
        a = sheet["A"+ str(i)].value
        b = sheet["B"+ str(i)].value
        c = sheet["C"+ str(i)].value
        tpl_data = (a, b, c)
        if a!= None or b != None or c != None:
            lst_data.append(tpl_data)
        workbook.close()
        return lst_data
